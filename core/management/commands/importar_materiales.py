from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from core.models import Material, Inventario


class Command(BaseCommand):
    help = "Importa materiales e inventario desde un archivo Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            "archivo",
            type=str,
            help="Ruta del archivo Excel"
        )

    def handle(self, *args, **options):

        archivo = options["archivo"]

        try:
            libro = load_workbook(
                archivo,
                read_only=True,
                data_only=True
            )
        except FileNotFoundError:
            self.stdout.write(
                self.style.ERROR(
                    f"No se encontró el archivo: {archivo}"
                )
            )
            return

        hoja = libro.active

        materiales = []

        # Saltamos la primera fila porque contiene los encabezados
        for fila in hoja.iter_rows(
            min_row=2,
            values_only=True
        ):
            item = fila[0]
            descripcion = fila[1]
            unidad = fila[2]

            # Ignorar filas sin ITEM o sin descripción
            if item is None or descripcion is None:
                continue

            materiales.append({
                "item": int(item),
                "descripcion": str(descripcion).strip(),
                "unidad": (
                    str(unidad).strip()
                    if unidad is not None
                    else None
                )
            })

        self.stdout.write(
            self.style.SUCCESS(
                f"Se encontraron {len(materiales)} materiales válidos."
            )
        )

        if not materiales:
            self.stdout.write(
                self.style.ERROR(
                    "No se encontraron materiales para importar."
                )
            )
            return

        # Mostrar algunos ejemplos
        self.stdout.write("\nPrimeros materiales encontrados:")

        for material in materiales[:5]:
            self.stdout.write(
                f'{material["item"]} | '
                f'{material["descripcion"]} | '
                f'{material["unidad"]}'
            )

        self.stdout.write("\nÚltimos materiales encontrados:")

        for material in materiales[-5:]:
            self.stdout.write(
                f'{material["item"]} | '
                f'{material["descripcion"]} | '
                f'{material["unidad"]}'
            )

        confirmar = input(
            "\n¿Desea importar estos materiales? [s/N]: "
        ).strip().lower()

        if confirmar != "s":
            self.stdout.write(
                self.style.WARNING(
                    "Importación cancelada."
                )
            )
            return

        creados_material = 0
        creados_inventario = 0

        for datos in materiales:

            material, creado = Material.objects.get_or_create(
                item=datos["item"],
                defaults={
                    "descripcion": datos["descripcion"],
                    "unidad": datos["unidad"],
                }
            )

            if creado:
                creados_material += 1

            inventario, creado = Inventario.objects.get_or_create(
                material=material,
                defaults={
                    "cantidad": 0
                }
            )

            if creado:
                creados_inventario += 1

        self.stdout.write(
            self.style.SUCCESS(
                "\nImportación terminada correctamente."
            )
        )

        self.stdout.write(
            f"Materiales creados: {creados_material}"
        )

        self.stdout.write(
            f"Registros de inventario creados: {creados_inventario}"
        )